"""Geração da planilha XLSX com as classificações (e feedback humano).

O frontend envia as linhas da tabela — incluindo a avaliação do usuário
(correto/incorreto) e o rótulo correto quando houver — e este módulo monta um
arquivo ``.xlsx`` com cabeçalho estilizado. O nome do arquivo carrega um
timestamp, pois espera-se salvar várias planilhas ao longo do uso.
"""

from __future__ import annotations

from datetime import datetime
from io import BytesIO

from app.schemas import ExportRow

# Ordem e título das colunas da planilha (pt-BR).
_COLUMNS: list[tuple[str, str]] = [
    ("document", "Documento"),
    ("predicted_label", "Rótulo predito"),
    ("confidence", "Confiança (%)"),
    ("threshold", "Limiar (%)"),
    ("needs_manual_review", "Revisar manual"),
    ("decision_label", "Decisão"),
    ("source", "Origem"),
    ("ocr_used", "OCR"),
    ("char_count", "Caracteres"),
    ("model_id", "Modelo"),
    ("feedback_status", "Avaliação"),
    ("correct_label", "Rótulo correto"),
    ("error", "Erro"),
]

_FEEDBACK_LABELS = {"correto": "Correto", "incorreto": "Incorreto"}


def export_filename(prefix: str = "classificacoes", now: datetime | None = None) -> str:
    """Nome de arquivo com timestamp, ex.: ``classificacoes_20260616_142233.xlsx``."""
    stamp = (now or datetime.now()).strftime("%Y%m%d_%H%M%S")
    return f"{prefix}_{stamp}.xlsx"


def _cell_value(row: ExportRow, key: str):
    value = getattr(row, key)
    if key in {"confidence", "threshold"}:
        # Percentual com uma casa (mantém número para permitir cálculos no Excel).
        return round(float(value) * 100, 1)
    if key == "needs_manual_review":
        return "Sim" if value else "Não"
    if key == "ocr_used":
        return "Sim" if value else "Não"
    if key == "feedback_status":
        return _FEEDBACK_LABELS.get(value or "", "")
    return "" if value is None else value


def build_xlsx(rows: list[ExportRow]) -> bytes:
    """Monta a planilha em memória e devolve os bytes do arquivo ``.xlsx``."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Classificações"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="9B1C2E")  # vermelho PGE
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, (_, title) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, (key, _) in enumerate(_COLUMNS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=_cell_value(row, key))

    # Larguras de coluna aproximadas para legibilidade.
    widths = [34, 40, 13, 11, 14, 40, 14, 7, 12, 22, 12, 40, 30]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_COLUMNS))}1"

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


__all__ = ["build_xlsx", "export_filename"]
