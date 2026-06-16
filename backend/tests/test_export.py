"""Testes do serviço de exportação XLSX."""

from __future__ import annotations

import io
from datetime import datetime

from openpyxl import load_workbook

from app.schemas import ExportRow
from app.services.export import build_xlsx, export_filename


def _rows() -> list[ExportRow]:
    return [
        ExportRow(
            document="peticao1.pdf",
            predicted_label="ICMS Declarado",
            confidence=0.912,
            threshold=0.5,
            needs_manual_review=False,
            decision_label="ICMS Declarado",
            source="pdf",
            ocr_used=False,
            char_count=1200,
            model_id="pge-tfidf-xgboost-v1",
            feedback_status="incorreto",
            correct_label="Servidor",
        ),
        ExportRow(
            document="peticao2.pdf",
            predicted_label="Servidor",
            confidence=0.40,
            threshold=0.5,
            needs_manual_review=True,
            decision_label="Revisar manualmente",
        ),
    ]


def test_export_filename_has_timestamp():
    name = export_filename(now=datetime(2026, 6, 16, 14, 22, 33))
    assert name == "classificacoes_20260616_142233.xlsx"


def test_build_xlsx_is_valid_workbook_with_header_and_rows():
    data = build_xlsx(_rows())
    wb = load_workbook(io.BytesIO(data))
    ws = wb.active
    assert ws.title == "Classificações"

    header = [c.value for c in ws[1]]
    assert header[0] == "Documento"
    assert "Rótulo predito" in header
    assert "Avaliação" in header
    assert "Rótulo correto" in header

    # Linha 1 de dados (linha 2 da planilha).
    assert ws.cell(row=2, column=1).value == "peticao1.pdf"
    # Confiança convertida para percentual com 1 casa.
    conf_col = header.index("Confiança (%)") + 1
    assert ws.cell(row=2, column=conf_col).value == 91.2
    # Booleanos viram Sim/Não.
    review_col = header.index("Revisar manual") + 1
    assert ws.cell(row=2, column=review_col).value == "Não"
    assert ws.cell(row=3, column=review_col).value == "Sim"
    # Feedback traduzido.
    fb_col = header.index("Avaliação") + 1
    assert ws.cell(row=2, column=fb_col).value == "Incorreto"
    # Sem avaliação → célula vazia (openpyxl relê string vazia como None).
    assert not ws.cell(row=3, column=fb_col).value


def test_build_xlsx_handles_empty_rows():
    data = build_xlsx([])
    wb = load_workbook(io.BytesIO(data))
    ws = wb.active
    assert [c.value for c in ws[1]][0] == "Documento"
    assert ws.max_row == 1
